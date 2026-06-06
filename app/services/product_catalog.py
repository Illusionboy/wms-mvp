import io
from dataclasses import dataclass

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product


@dataclass(frozen=True)
class _CatalogRow:
    jan_code: str
    name_jp: str
    name_zh: str | None
    units_per_case: int | None


async def import_product_catalog_from_bytes(
    session: AsyncSession,
    content: bytes,
) -> dict[str, int]:
    """Upsert products from an Excel catalog file (bytes).

    Returns counts: created, updated, skipped.
    """
    rows = _parse_catalog_bytes(content)
    created = updated = 0

    for row in rows:
        product = await session.get(Product, row.jan_code)
        if product is None:
            session.add(Product(
                jan_code=row.jan_code,
                name_jp=row.name_jp,
                name_zh=row.name_zh,
                units_per_case=row.units_per_case,
            ))
            created += 1
            continue

        changed = False
        if row.name_jp and product.name_jp != row.name_jp:
            product.name_jp = row.name_jp
            changed = True
        if row.name_zh and product.name_zh != row.name_zh:
            product.name_zh = row.name_zh
            changed = True
        if row.units_per_case is not None and product.units_per_case != row.units_per_case:
            product.units_per_case = row.units_per_case
            changed = True
        if changed:
            updated += 1

    await session.commit()
    return {"created": created, "updated": updated, "skipped": len(rows) - created - updated}


def _parse_catalog_bytes(content: bytes) -> list[_CatalogRow]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header is None:
        raise ValueError("Excel 文件为空")
    col = _detect_columns(header)

    rows_by_jan: dict[str, _CatalogRow] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        jan = _normalize_jan(row[col["jan"]] if len(row) > col["jan"] else None)
        if not jan:
            continue
        name_jp = _clean(row[col["name_jp"]] if len(row) > col["name_jp"] else None) or jan
        name_zh = _clean(row[col["name_zh"]] if len(row) > col["name_zh"] else None) or None
        upc = _positive_int(row[col["units_per_case"]] if len(row) > col["units_per_case"] else None)
        rows_by_jan[jan] = _CatalogRow(
            jan_code=jan,
            name_jp=name_jp[:255],
            name_zh=name_zh[:255] if name_zh else None,
            units_per_case=upc,
        )
    wb.close()
    if not rows_by_jan:
        raise ValueError("未找到有效数据行，请确认文件包含 JAN 列和商品名列")
    return list(rows_by_jan.values())


def _detect_columns(header: tuple) -> dict[str, int]:
    normalized = {_clean(v): i for i, v in enumerate(header)}
    specs: dict[str, list[str]] = {
        "jan": ["JAN", "jan", "条码", "barcode"],
        "name_jp": ["商品名", "日文", "name_jp"],
        "name_zh": ["中文", "name_zh"],
        "units_per_case": ["箱入れ数", "箱入数", "箱入", "units_per_case"],
    }
    col: dict[str, int] = {}
    for key, candidates in specs.items():
        for name in candidates:
            if name in normalized:
                col[key] = normalized[name]
                break
        if key not in col:
            if key in ("name_zh", "units_per_case"):
                col[key] = 9999  # optional — safe sentinel, len check will skip
            else:
                raise ValueError(f"Excel 缺少必要列: {candidates[0]}（也可以是 {candidates}）")
    return col


def _normalize_jan(value: object) -> str:
    text = _clean(value)
    return "".join(ch for ch in text if ch.isdigit())


def _positive_int(value: object) -> int | None:
    text = _clean(value)
    if not text:
        return None
    try:
        n = int(float(text))
    except ValueError:
        return None
    return n if n > 0 else None


def _clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in ("nan", "none", "null"):
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return text
