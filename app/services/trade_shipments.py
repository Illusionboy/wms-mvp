import asyncio
import json
from io import BytesIO
from typing import NamedTuple

import openpyxl
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.inventory_record import InventoryRecord
from app.models.product import Product
from app.models.stock_transaction import StockTransaction
from app.models.trade_shipment_draft import TradeShipmentDraft
from app.schemas.inventory import (
    ProductRead,
    StockOutCreate,
    StockTransactionRead,
    TradeShipmentDraftDocument,
    TradeShipmentDraftPreview,
    TradeShipmentImportResult,
    TradeShipmentIssue,
    TradeShipmentLine,
    TradeShipmentMutation,
)
from app.services.inventory import (
    InsufficientStockError,
    InventoryRecordNotFoundError,
    InventoryTargetNotFoundError,
    resolve_warehouse,
    search_inventory_items,
    stock_out_item,
)


DEFAULT_TRADE_WAREHOUSE = "普通仓库"
TRADE_SHIPMENT_SOURCE = "trade_shipment"

_HEADER_ALIASES = {
    "jan_code": ("jan", "jan码", "jan码", "条码", "barcode"),
    "product_name": ("品名", "商品名", "名称"),
    "box_count": ("箱数",),
    "units_per_box": ("箱入",),
    "quantity": ("数量", "qty", "quantity", "総数", "总数"),
}


class _ValidationResult(NamedTuple):
    blocking_issues: list[TradeShipmentIssue]
    needs_decision: list[TradeShipmentIssue]
    resolved: list[tuple[int, str, InventoryRecord]]


class _TradeShipmentLinesSchema(BaseModel):
    lines: list[TradeShipmentLine] = Field(default_factory=list)


def parse_trade_shipment_excel(content: bytes) -> list[TradeShipmentLine]:
    """Each worksheet name is a customer code (mm/kk/cp/...). Columns jan/品名/箱数/箱入
    are detected from the header row, in any order.

    Real shipment sheets routinely merge the jan/品名 cells vertically when the same
    JAN is split across several box-size rows (e.g. 4箱×102入 + 12箱×48入 + 1箱×16入 for
    one JAN) — openpyxl leaves every non-anchor cell in a merged range as None, so without
    backfilling those rows look JAN-less and get silently dropped, undercounting the total.
    """
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    lines: list[TradeShipmentLine] = []
    for sheet in workbook.worksheets:
        customer_name = sheet.title.strip()
        if not customer_name:
            continue
        rows = list(sheet.iter_rows(values_only=True))
        column_index = _detect_columns(rows)
        if column_index is None:
            continue
        header_row_index, columns = column_index

        merged_value_map: dict[tuple[int, int], object] = {}
        for merged_range in sheet.merged_cells.ranges:
            anchor_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_value_map[(r, c)] = anchor_value

        for offset, row in enumerate(rows[header_row_index + 1:]):
            row_number = header_row_index + 2 + offset  # openpyxl rows are 1-indexed
            line = _parse_trade_row(row, columns, customer_name, merged_value_map, row_number)
            if line is not None:
                lines.append(line)
    return lines


def _cell_value(row: tuple, col: int | None, merged_value_map: dict[tuple[int, int], object], row_number: int) -> object:
    if col is None or col >= len(row):
        return None
    value = row[col]
    if value is None:
        value = merged_value_map.get((row_number, col + 1))
    return value


def _detect_columns(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    for row_index, row in enumerate(rows):
        columns: dict[str, int] = {}
        for col_index, cell in enumerate(row):
            text = _clean_text(cell).lower()
            if not text:
                continue
            for field, aliases in _HEADER_ALIASES.items():
                if field in columns:
                    continue
                if any(alias in text for alias in aliases):
                    columns[field] = col_index
        if {"jan_code", "box_count", "units_per_box"}.issubset(columns):
            return row_index, columns
    return None


def _parse_trade_row(
    row: tuple,
    columns: dict[str, int],
    customer_name: str,
    merged_value_map: dict[tuple[int, int], object],
    row_number: int,
) -> TradeShipmentLine | None:
    raw_jan = _cell_value(row, columns.get("jan_code"), merged_value_map, row_number)
    jan_digits = "".join(ch for ch in _clean_text(raw_jan) if ch.isdigit())
    if not jan_digits:
        return None
    box_count = _parse_positive_int(_cell_value(row, columns.get("box_count"), merged_value_map, row_number))
    units_per_box = _parse_positive_int(_cell_value(row, columns.get("units_per_box"), merged_value_map, row_number))
    quantity_override = _parse_positive_int(_cell_value(row, columns.get("quantity"), merged_value_map, row_number))

    if box_count is None or units_per_box is None:
        # 箱数/箱入缺一个但有现成的总数/QTY列（例如合箱临时凑数的尾行）：
        # 折算成 1箱×总数件，保留正确的总量，而不是直接丢掉这一行。
        if quantity_override is None:
            return None
        box_count, units_per_box, quantity = 1, quantity_override, quantity_override
    else:
        quantity = box_count * units_per_box

    product_name = None
    name_value = _cell_value(row, columns.get("product_name"), merged_value_map, row_number)
    if name_value is not None:
        product_name = _clean_text(name_value) or None
    return TradeShipmentLine(
        customer_name=customer_name,
        jan_code=jan_digits,
        product_name=product_name,
        box_count=box_count,
        units_per_box=units_per_box,
        quantity=quantity,
    )


_TRADE_SHIPMENT_IMAGE_PROMPT = """
你是 WMS 贸易出库单识别助手。图片中是手写/打印的批发出货单照片，请将其整理成严格 JSON。

表格结构说明：
- 一张图片中可能包含多个区块，每个区块对应一个客户代码（手写圈出的字母组合，如 mm / kk / cp / hn / xm / mmm），
  该代码标注在区块附近，请将其原样转为小写填入 customer_name。
- 每个区块是一个表格，列包括：jan码、品名（日文商品名）、箱数、箱入。列的顺序可能与示例不同，请按表头识别。
- jan码列有时含有非数字前缀（例如"外箱jan4902806131311"），请只提取其中的数字序列作为 jan_code。
- 数量计算：quantity = 箱数 × 箱入（合箱时箱数=1，箱入即为该行实际数量，公式同样适用，无需特殊处理）。

输出规则：
- 为每一行商品输出一个 line：customer_name、jan_code、product_name、box_count、units_per_box、quantity。
- 无法辨认 jan码或数量的行不要输出。
- 不要输出解释，只输出符合 schema 的 JSON。
""".strip()


def _parse_trade_shipment_images_sync(images: list[tuple[bytes, str]]) -> list[TradeShipmentLine]:
    from google import genai
    from google.genai import types

    if not settings.gemini_api_key:
        raise RuntimeError("GEMINI_API_KEY 未配置，贸易出库拍照识别功能不可用。")
    client = genai.Client(api_key=settings.gemini_api_key)
    parts = [types.Part.from_bytes(data=data, mime_type=mime) for data, mime in images]
    parts.append(types.Part.from_text(text=_TRADE_SHIPMENT_IMAGE_PROMPT))
    response = client.models.generate_content(
        model=settings.gemini_model,
        contents=parts,
        config={
            "response_mime_type": "application/json",
            "response_json_schema": _TradeShipmentLinesSchema.model_json_schema(),
        },
    )
    raw_text = response.text or "{}"
    data = json.loads(raw_text)
    return _TradeShipmentLinesSchema.model_validate(data).lines


async def parse_trade_shipment_images_with_gemini(images: list[tuple[bytes, str]]) -> list[TradeShipmentLine]:
    return await asyncio.to_thread(_parse_trade_shipment_images_sync, images)


async def create_trade_shipment_draft(
    session: AsyncSession,
    lines: list[TradeShipmentLine],
    original_filename: str,
    warehouse_name: str = DEFAULT_TRADE_WAREHOUSE,
) -> TradeShipmentDraft:
    document = TradeShipmentDraftDocument(warehouse_name=warehouse_name, lines=lines)
    draft = TradeShipmentDraft(
        original_filename=original_filename,
        status="parsed",
        warehouse_name=warehouse_name,
        document=document.model_dump(mode="json"),
    )
    session.add(draft)
    await session.commit()
    await session.refresh(draft)
    return draft


async def get_trade_shipment_draft(
    session: AsyncSession,
    draft_id: int,
    *,
    with_for_update: bool = False,
) -> TradeShipmentDraft | None:
    stmt = select(TradeShipmentDraft).where(TradeShipmentDraft.id == draft_id)
    if with_for_update:
        stmt = stmt.with_for_update()
    return await session.scalar(stmt)


async def update_trade_shipment_draft_lines(
    session: AsyncSession,
    draft: TradeShipmentDraft,
    lines: list[TradeShipmentLine],
) -> TradeShipmentDraft:
    document = TradeShipmentDraftDocument.model_validate(draft.document)
    document = document.model_copy(update={"lines": lines})
    draft.document = document.model_dump(mode="json")
    await session.commit()
    await session.refresh(draft)
    return draft


async def preview_trade_shipment_draft(
    session: AsyncSession,
    draft: TradeShipmentDraft,
) -> TradeShipmentDraftPreview:
    document = TradeShipmentDraftDocument.model_validate(draft.document)
    vr = await _validate_trade_shipment_lines(
        session=session,
        lines=document.lines,
        warehouse_name=document.warehouse_name,
    )
    return TradeShipmentDraftPreview(
        draft_id=draft.id,
        total_lines=len(document.lines),
        ok_count=len(vr.resolved),
        needs_decision=vr.needs_decision,
        blocking_issues=vr.blocking_issues,
        document=document,
    )


async def apply_trade_shipment_draft(
    session: AsyncSession,
    draft: TradeShipmentDraft,
    force_negative_jans: set[str] | None = None,
    user_id: int | None = None,
) -> TradeShipmentImportResult:
    document = TradeShipmentDraftDocument.model_validate(draft.document)
    vr = await _validate_trade_shipment_lines(
        session=session,
        lines=document.lines,
        warehouse_name=document.warehouse_name,
    )

    if vr.blocking_issues:
        return TradeShipmentImportResult(
            applied=False,
            total_lines=len(document.lines),
            issues=vr.blocking_issues,
        )

    force_negative_jans = force_negative_jans or set()
    mutations: list[TradeShipmentMutation] = []
    skipped_duplicates = 0
    force_negated_count = 0
    remaining_issues: list[TradeShipmentIssue] = []

    for line_index, jan_code, record in vr.resolved:
        line = document.lines[line_index]
        reference_id = f"trade:{draft.id}:{line_index}"
        existing = await session.scalar(
            select(StockTransaction.id).where(
                StockTransaction.source == TRADE_SHIPMENT_SOURCE,
                StockTransaction.reference_id == reference_id,
            ).limit(1)
        )
        if existing:
            skipped_duplicates += 1
            continue
        try:
            result = await stock_out_item(
                session=session,
                payload=StockOutCreate(
                    sku=jan_code,
                    warehouse_id=record.warehouse_id,
                    quantity=line.quantity,
                    source=TRADE_SHIPMENT_SOURCE,
                    location_code=record.location_code,
                    expiration_date=record.expiration_date,
                    reference_id=reference_id,
                    note=f"贸易出库 客户:{line.customer_name}",
                    customer=line.customer_name,
                ),
                commit=False,
                user_id=user_id,
            )
        except (InsufficientStockError, InventoryRecordNotFoundError) as exc:
            raise RuntimeError(f"Trade shipment apply failed after validation: {exc}") from exc
        mutations.append(
            TradeShipmentMutation(
                jan_code=jan_code,
                customer_name=line.customer_name,
                quantity=line.quantity,
                transaction=StockTransactionRead.model_validate(result.transaction),
                low_stock_alert=result.low_stock_alert,
            )
        )

    warehouse = await resolve_warehouse(session, document.warehouse_name)
    for issue in vr.needs_decision:
        if issue.jan_code not in force_negative_jans or warehouse is None:
            remaining_issues.append(issue)
            continue
        line = document.lines[issue.line_index]
        reference_id = f"trade:{draft.id}:{issue.line_index}"
        existing = await session.scalar(
            select(StockTransaction.id).where(
                StockTransaction.source == TRADE_SHIPMENT_SOURCE,
                StockTransaction.reference_id == reference_id,
            ).limit(1)
        )
        if existing:
            skipped_duplicates += 1
            continue
        try:
            result = await stock_out_item(
                session=session,
                payload=StockOutCreate(
                    sku=issue.jan_code,
                    warehouse_id=warehouse.id,
                    quantity=line.quantity,
                    source=TRADE_SHIPMENT_SOURCE,
                    reference_id=reference_id,
                    note=f"贸易出库 客户:{line.customer_name} (负库存确认)",
                    customer=line.customer_name,
                ),
                commit=False,
                user_id=user_id,
                force_negative=True,
            )
        except InventoryTargetNotFoundError:
            remaining_issues.append(issue)
            continue
        except (InsufficientStockError, InventoryRecordNotFoundError) as exc:
            raise RuntimeError(f"Trade shipment force-negative apply failed: {exc}") from exc
        mutations.append(
            TradeShipmentMutation(
                jan_code=issue.jan_code,
                customer_name=line.customer_name,
                quantity=line.quantity,
                transaction=StockTransactionRead.model_validate(result.transaction),
                low_stock_alert=result.low_stock_alert,
            )
        )
        force_negated_count += 1

    has_skips = bool(skipped_duplicates or remaining_issues)
    draft.status = "applied_with_skips" if has_skips else "applied"
    await session.commit()
    await session.refresh(draft)

    return TradeShipmentImportResult(
        applied=True,
        total_lines=len(document.lines),
        mutations=mutations,
        issues=remaining_issues,
        skipped_duplicates=skipped_duplicates,
        force_negated_count=force_negated_count,
    )


async def _validate_trade_shipment_lines(
    session: AsyncSession,
    lines: list[TradeShipmentLine],
    warehouse_name: str,
) -> _ValidationResult:
    warehouse = await resolve_warehouse(session, warehouse_name)
    blocking_issues: list[TradeShipmentIssue] = []
    needs_decision: list[TradeShipmentIssue] = []
    resolved: list[tuple[int, str, InventoryRecord]] = []

    if not lines:
        blocking_issues.append(TradeShipmentIssue(
            line_index=-1, jan_code="", issue_type="empty_document",
            message="未解析到任何商品行。",
        ))
    if warehouse is None:
        blocking_issues.append(TradeShipmentIssue(
            line_index=-1, jan_code="", issue_type="warehouse_not_found",
            message=f"未找到仓库: {warehouse_name}",
        ))

    for index, line in enumerate(lines):
        products = await search_inventory_items(session=session, keyword=line.jan_code, limit=6)
        if not products:
            needs_decision.append(TradeShipmentIssue(
                line_index=index,
                jan_code=line.jan_code,
                issue_type="product_not_found",
                message=f"未找到商品: {line.jan_code}，请在预览中修正 JAN 码",
                quantity_needed=line.quantity,
            ))
            continue
        if len(products) > 1:
            blocking_issues.append(TradeShipmentIssue(
                line_index=index,
                jan_code=line.jan_code,
                issue_type="ambiguous_product",
                message="匹配到多个商品，请使用完整 JAN 码确认。",
                candidates=[ProductRead.model_validate(p) for p in products],
            ))
            continue
        if warehouse is None:
            continue

        jan_code = products[0].jan_code
        record = await _resolve_first_inventory_record(
            session=session,
            jan_code=jan_code,
            warehouse_id=warehouse.id,
        )
        if record is None:
            needs_decision.append(TradeShipmentIssue(
                line_index=index,
                jan_code=jan_code,
                issue_type="inventory_record_not_found",
                message=f"{warehouse_name} 无 {jan_code} 库存记录",
                current_stock=None,
                quantity_needed=line.quantity,
            ))
            continue
        if record.quantity < line.quantity:
            needs_decision.append(TradeShipmentIssue(
                line_index=index,
                jan_code=jan_code,
                issue_type="insufficient_stock",
                message=f"库存不足：{jan_code} 现有 {record.quantity}，出库需 {line.quantity}",
                current_stock=record.quantity,
                quantity_needed=line.quantity,
            ))
            continue
        resolved.append((index, jan_code, record))

    return _ValidationResult(blocking_issues, needs_decision, resolved)


async def _resolve_first_inventory_record(
    session: AsyncSession,
    jan_code: str,
    warehouse_id: int,
) -> InventoryRecord | None:
    return await session.scalar(
        select(InventoryRecord)
        .where(
            InventoryRecord.product_jan == jan_code,
            InventoryRecord.warehouse_id == warehouse_id,
        )
        .order_by(InventoryRecord.id.asc())
        .limit(1)
    )


def _parse_positive_int(value: object) -> int | None:
    text = _clean_text(value)
    if not text:
        return None
    try:
        parsed = int(float(text))
    except ValueError:
        return None
    if parsed <= 0:
        return None
    return parsed


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return text
