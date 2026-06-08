import hashlib
import io
import json
import logging
import re
from datetime import date, datetime, time, timezone
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup, Tag
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.inventory_count_draft import InventoryCountDraft
from app.models.inventory_record import InventoryRecord
from app.models.product import Product
from app.models.stock_transaction import StockTransaction, StockTransactionType
from app.schemas.inventory_count import (
    CountDraftLine,
    InventoryCountApplyResult,
    InventoryCountDocument,
    InventoryCountDraftRead,
    QinsiColumnMap,
    QinsiSession,
    QinsiSessionListResult,
)
from app.services.inventory import resolve_customer, resolve_warehouse

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column map cache (file-backed, survives restarts)
# ---------------------------------------------------------------------------

_CACHE_PATH = Path(__file__).parent.parent / "data" / "qinsi_schema_cache.json"

# Hardcoded fallback indices (秦丝生意通 list5 as of 2024)
_HARDCODED_MAP = QinsiColumnMap(jan_col=7, name_col=3, count_col=12, source="hardcoded")
# list4 (明细表) column map — used when list5 is empty (JS not rendered on save)
_LIST4_MAP = QinsiColumnMap(jan_col=8, name_col=3, count_col=22, source="hardcoded")

# Keyword hints for each column type
_COL_HINTS: dict[str, list[str]] = {
    "jan":   ["条码", "barcode", "货号", "sku", "商品编码"],
    "name":  ["商品名称", "品名", "名称", "商品"],
    "count": ["盘点数量", "盘点", "实盘", "数量", "count"],
}


def _load_cache() -> dict[str, QinsiColumnMap]:
    try:
        if _CACHE_PATH.exists():
            data = json.loads(_CACHE_PATH.read_text())
            return {k: QinsiColumnMap(**v) for k, v in data.items()}
    except Exception:
        pass
    return {}


def _save_cache(cache: dict[str, QinsiColumnMap]) -> None:
    try:
        _CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        _CACHE_PATH.write_text(json.dumps({k: v.model_dump() for k, v in cache.items()}, ensure_ascii=False, indent=2))
    except Exception as exc:
        logger.warning("Failed to save qinsi schema cache: %s", exc)


def _header_fingerprint(headers: list[str]) -> str:
    return hashlib.md5("|".join(headers).encode()).hexdigest()[:12]


# ---------------------------------------------------------------------------
# Column detection — layer ① keyword matching
# ---------------------------------------------------------------------------

def _detect_columns_from_headers(headers: list[str]) -> QinsiColumnMap | None:
    """Try to identify column indices by keyword matching against header cells."""
    found: dict[str, int] = {}
    for idx, cell in enumerate(headers):
        cell_lower = cell.lower()
        for col_type, hints in _COL_HINTS.items():
            if col_type not in found and any(h in cell_lower for h in hints):
                found[col_type] = idx
    if "jan" in found and "count" in found:
        return QinsiColumnMap(
            jan_col=found["jan"],
            name_col=found.get("name", 3),
            count_col=found["count"],
            source="detected",
        )
    return None


# ---------------------------------------------------------------------------
# Column detection — layer ③ LLM repair
# ---------------------------------------------------------------------------

async def _llm_repair_column_map(headers: list[str], sample_rows: list[list[str]]) -> QinsiColumnMap:
    """Ask Gemini to infer column indices when automatic detection fails."""
    from google import genai  # type: ignore
    from app.core.config import settings

    client = genai.Client(api_key=settings.gemini_api_key)

    headers_str = json.dumps(headers, ensure_ascii=False)
    samples_str = "\n".join(json.dumps(r, ensure_ascii=False) for r in sample_rows[:3])

    prompt = f"""
你是数据解析专家。以下是一个日本商品管理系统（秦丝生意通）导出的盘点HTML表格的列标题：
{headers_str}

以下是前3行数据示例：
{samples_str}

请找出以下三列分别对应哪个索引（从0开始）：
1. jan_col: JAN条码/商品条码（13位数字，或含数字的字段）
2. name_col: 商品名称（日文或中文商品名）
3. count_col: 盘点数量（整数，表示实际清点数量）

只返回如下JSON，不要其他内容：
{{"jan_col": <int>, "name_col": <int>, "count_col": <int>}}
"""
    response = client.models.generate_content(
        model=settings.gemini_model,
        contents=prompt,
    )
    text = response.text.strip()
    # Extract JSON from response
    match = re.search(r'\{[^}]+\}', text)
    if not match:
        raise ValueError(f"LLM 未能返回有效的列映射 JSON，原始响应: {text[:200]}")
    data = json.loads(match.group())
    return QinsiColumnMap(
        jan_col=int(data["jan_col"]),
        name_col=int(data["name_col"]),
        count_col=int(data["count_col"]),
        source="llm_repaired",
    )


# ---------------------------------------------------------------------------
# Core HTML table parser (shared by list-sessions + upload)
# ---------------------------------------------------------------------------

def _find_count_tables(soup: BeautifulSoup) -> list[Tag]:
    """Find all 秦丝盘点 summary tables in the document.

    Looks for tables whose id starts with 'list5'. A single HTML may contain
    multiple sessions (e.g. list5, list5_2) if the user exported several
    count batches together.
    """
    tables = soup.find_all("table", id=re.compile(r"^list5"))
    return [t for t in tables if isinstance(t, Tag)]


def _extract_table_rows(table: Tag) -> tuple[list[str], list[list[str]]]:
    """Return (header_cells, data_rows) from a <table> element."""
    all_rows = table.find_all("tr")
    header: list[str] = []
    data_rows: list[list[str]] = []
    for tr in all_rows:
        cells = [td.get_text(strip=True) for td in tr.find_all("td")]
        if not cells:
            continue
        if not header and not cells[0].isdigit():
            # First non-data row is the header
            header = cells
        elif cells[0].isdigit():
            data_rows.append(cells)
    return header, data_rows


def _date_hint_near_table(table: Tag) -> str | None:
    """Look for a date string in the preceding sibling elements."""
    date_re = re.compile(r'\d{4}[-/年]\d{1,2}[-/月]\d{1,2}')
    for sibling in table.previous_siblings:
        if isinstance(sibling, Tag):
            text = sibling.get_text(" ", strip=True)
            m = date_re.search(text)
            if m:
                return m.group()
    return None


def _title_hint_near_table(table: Tag) -> str | None:
    """Look for a title heading near the table."""
    for sibling in table.previous_siblings:
        if isinstance(sibling, Tag) and sibling.name in ("h1","h2","h3","h4","div","p"):
            text = sibling.get_text(" ", strip=True)
            if text:
                return text[:80]
    return None


async def _resolve_column_map(
    headers: list[str],
    sample_rows: list[list[str]],
) -> QinsiColumnMap:
    """Three-layer column detection. Returns a QinsiColumnMap."""
    fingerprint = _header_fingerprint(headers)

    # Layer ② cache
    cache = _load_cache()
    if fingerprint in cache:
        cached = cache[fingerprint]
        logger.info("Qinsi schema cache hit: fingerprint=%s map=%s", fingerprint, cached)
        return cached.model_copy(update={"source": "cached"})

    # Layer ① keyword detection
    detected = _detect_columns_from_headers(headers)
    if detected is not None:
        logger.info("Qinsi schema detected from headers: %s", detected)
        cache[fingerprint] = detected
        _save_cache(cache)
        return detected

    # Layer ③ LLM repair
    logger.warning(
        "Qinsi schema detection failed for headers=%s — calling Gemini", headers
    )
    repaired = await _llm_repair_column_map(headers, sample_rows)
    logger.info("Qinsi schema repaired by LLM: %s", repaired)
    cache[fingerprint] = repaired
    _save_cache(cache)
    return repaired


def _apply_column_map(
    data_rows: list[list[str]],
    col_map: QinsiColumnMap,
) -> list[tuple[str, str, int]]:
    """Extract (jan, name, qty) tuples using a resolved column map."""
    result: list[tuple[str, str, int]] = []
    for cells in data_rows:
        raw_jan = cells[col_map.jan_col].strip() if len(cells) > col_map.jan_col else ""
        jan = "".join(ch for ch in raw_jan if ch.isdigit())
        if not jan:
            continue
        name = cells[col_map.name_col] if len(cells) > col_map.name_col else ""
        try:
            qty = int(float(cells[col_map.count_col])) if len(cells) > col_map.count_col else 0
        except ValueError:
            continue
        result.append((jan, name, max(qty, 0)))
    return result


# ---------------------------------------------------------------------------
# Public: list sessions (new endpoint)
# ---------------------------------------------------------------------------

def list_qinsi_sessions(content: bytes, filename: str) -> QinsiSessionListResult:
    """Parse an HTML file and return a list of count sessions without full processing."""
    text = content.decode("utf-8", errors="replace")
    soup = BeautifulSoup(text, "lxml")
    tables = _find_count_tables(soup)
    if not tables:
        raise ValueError("不是有效的秦丝生意通盘点单 HTML 文件（找不到 list5 表格）")

    sessions: list[QinsiSession] = []
    for idx, table in enumerate(tables):
        _, data_rows = _extract_table_rows(table)
        sessions.append(QinsiSession(
            session_index=idx,
            table_id=table.get("id", f"list5_{idx}"),
            item_count=len(data_rows),
            date_hint=_date_hint_near_table(table),
            title_hint=_title_hint_near_table(table),
        ))
    return QinsiSessionListResult(
        session_count=len(sessions),
        sessions=sessions,
        filename=filename,
    )


# ---------------------------------------------------------------------------
# File parsers
# ---------------------------------------------------------------------------

async def parse_count_file(
    content: bytes,
    filename: str,
    session_index: int = 0,
) -> list[tuple[str, str, int]]:
    """Return list of (jan_code, product_name, count_quantity) from uploaded file.

    For HTML files, session_index selects which count batch to parse when the
    file contains multiple. Duplicate JAN codes are merged (summed).
    """
    lower = filename.lower()
    if lower.endswith(".html") or lower.endswith(".htm"):
        raw = await _parse_qinsi_html(content, session_index=session_index)
    elif lower.endswith(".xlsx") or lower.endswith(".xls"):
        raw = _parse_count_excel(content)
    else:
        raise ValueError(f"不支持的文件格式，请上传 .html 或 .xlsx 文件（收到: {filename}）")
    return _merge_duplicate_jans(raw)


def _merge_duplicate_jans(
    rows: list[tuple[str, str, int]],
) -> list[tuple[str, str, int]]:
    seen: dict[str, tuple[str, int]] = {}
    for jan, name, qty in rows:
        if jan in seen:
            existing_name, existing_qty = seen[jan]
            seen[jan] = (existing_name, existing_qty + qty)
        else:
            seen[jan] = (name, qty)
    return [(jan, name, qty) for jan, (name, qty) in seen.items()]


async def _parse_qinsi_html(
    content: bytes,
    session_index: int = 0,
) -> list[tuple[str, str, int]]:
    text = content.decode("utf-8", errors="replace")
    soup = BeautifulSoup(text, "lxml")
    tables = _find_count_tables(soup)
    if not tables:
        raise ValueError("不是有效的秦丝生意通盘点单 HTML 文件（找不到 list5 表格）")
    if session_index >= len(tables):
        raise ValueError(
            f"session_index={session_index} 超出范围，文件共 {len(tables)} 个盘点记录（0~{len(tables)-1}）"
        )
    table = tables[session_index]
    headers, data_rows = _extract_table_rows(table)

    if not data_rows:
        # list5 is empty (JS-rendered grid not captured on save); try list4 (明细表)
        list4 = soup.find("table", id="list4")
        if list4 and isinstance(list4, Tag):
            _, data_rows = _extract_table_rows(list4)
            if data_rows:
                logger.info("list5 为空，自动 fallback 到 list4 解析（使用固定列映射）")
                return _apply_column_map(data_rows, _LIST4_MAP)
        raise ValueError(f"第 {session_index} 个盘点记录中没有数据行")

    # If no header found, fall back to hardcoded map directly
    if not headers:
        logger.warning("No header row found in list5 table, using hardcoded column map")
        col_map = _HARDCODED_MAP
    else:
        col_map = await _resolve_column_map(headers, data_rows)

    rows = _apply_column_map(data_rows, col_map)
    if not rows:
        raise ValueError(
            f"HTML 解析结果为空（列映射: JAN={col_map.jan_col} 名称={col_map.name_col} "
            f"数量={col_map.count_col} 来源={col_map.source}）。"
            "请检查文件来自秦丝生意通盘点单，或重试。"
        )
    logger.info(
        "Parsed %d rows from session_index=%d using col_map source=%s",
        len(rows), session_index, col_map.source,
    )
    return rows


def _parse_count_excel(content: bytes) -> list[tuple[str, str, int]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # Detect header row — first match wins to avoid generic keywords overwriting specific ones
    jan_col = count_col = name_col = None
    header = next(rows_iter, None)
    if header is None:
        raise ValueError("Excel 文件为空")
    for i, cell in enumerate(header):
        val = str(cell or "").lower()
        if jan_col is None and any(k in val for k in ["jan", "条码", "barcode", "货号"]):
            jan_col = i
        # Use specific "盘点数量"/"盘点数"/"实盘" — avoid generic "数量" which also matches 盘点前数量/盈亏数量
        if count_col is None and any(k in val for k in ["盘点数量", "盘点数", "实盘", "count"]):
            count_col = i
        if name_col is None and any(k in val for k in ["商品", "名称", "name"]):
            name_col = i

    if jan_col is None or count_col is None:
        raise ValueError(
            f"Excel 缺少必要列（需要条码列和数量列）。"
            f"检测到的列: {list(header)}"
        )

    result: list[tuple[str, str, int]] = []
    for row in rows_iter:
        jan_raw = str(row[jan_col] or "").strip()
        # Strip Excel float suffix (.0) without removing legitimate trailing zeros from JAN
        jan = jan_raw.split(".")[0] if "." in jan_raw else jan_raw
        if not jan or not jan.isdigit():
            continue
        try:
            qty = int(float(str(row[count_col] or 0)))
        except (ValueError, TypeError):
            continue
        name = str(row[name_col] or "") if name_col is not None else jan
        result.append((jan, name, max(qty, 0)))
    if not result:
        raise ValueError("Excel 文件中未找到有效数据行")
    return result


# ---------------------------------------------------------------------------
# Draft creation
# ---------------------------------------------------------------------------

async def create_inventory_count_draft(
    session: AsyncSession,
    content: bytes,
    filename: str,
    count_date: date,
    warehouse_name: str,
    customer_name: str | None,
    session_index: int = 0,
    cover_uncovered: bool = True,
) -> InventoryCountDraft:
    parsed = await parse_count_file(content, filename, session_index=session_index)

    warehouse = await resolve_warehouse(session, warehouse_name)
    if warehouse is None:
        raise ValueError(f"仓库不存在: {warehouse_name}")
    customer = None
    if customer_name:
        customer = await resolve_customer(session, customer_name)
        if customer is None:
            raise ValueError(f"客户不存在: {customer_name}")

    lines = await _compute_draft_lines(
        session=session,
        parsed=parsed,
        count_date=count_date,
        warehouse_id=warehouse.id,
        customer_id=customer.id if customer else None,
        cover_uncovered=cover_uncovered,
    )

    document = InventoryCountDocument(
        count_date=count_date,
        warehouse_name=warehouse_name,
        customer_name=customer_name,
        lines=lines,
    )
    draft = InventoryCountDraft(
        original_filename=filename,
        count_date=count_date,
        warehouse_name=warehouse_name,
        customer_name=customer_name,
        status="parsed",
        document=document.model_dump(mode="json"),
    )
    session.add(draft)
    await session.commit()
    await session.refresh(draft)
    return draft


async def _compute_draft_lines(
    session: AsyncSession,
    parsed: list[tuple[str, str, int]],
    count_date: date,
    warehouse_id: int,
    customer_id: int | None,
    cover_uncovered: bool = True,
) -> list[CountDraftLine]:
    # Transactions strictly after the end of count_date
    cutoff = datetime.combine(count_date, time(23, 59, 59)).replace(tzinfo=timezone.utc)

    lines: list[CountDraftLine] = []
    covered_jans: set[str] = set()

    for jan_code, product_name, count_qty in parsed:
        covered_jans.add(jan_code)

        current_qty_result = await session.scalar(
            select(func.coalesce(func.sum(InventoryRecord.quantity), 0))
            .where(
                InventoryRecord.product_jan == jan_code,
                InventoryRecord.warehouse_id == warehouse_id,
                _customer_filter(customer_id),
            )
        )
        current_qty = int(current_qty_result or 0)

        has_record = bool(await session.scalar(
            select(InventoryRecord.id)
            .where(
                InventoryRecord.product_jan == jan_code,
                InventoryRecord.warehouse_id == warehouse_id,
                _customer_filter(customer_id),
            )
            .limit(1)
        ))

        known_product = bool(await session.scalar(
            select(Product.jan_code).where(Product.jan_code == jan_code).limit(1)
        ))

        delta_result = await session.scalar(
            select(func.coalesce(func.sum(StockTransaction.quantity_change), 0))
            .join(InventoryRecord, StockTransaction.inventory_record_id == InventoryRecord.id)
            .where(
                InventoryRecord.product_jan == jan_code,
                InventoryRecord.warehouse_id == warehouse_id,
                StockTransaction.created_at > cutoff,
            )
        )
        delta = int(delta_result or 0)

        target_qty = count_qty + delta
        lines.append(
            CountDraftLine(
                jan_code=jan_code,
                product_name=product_name,
                count_quantity=count_qty,
                delta_after_count=delta,
                target_quantity=target_qty,
                current_quantity=current_qty,
                adjust_delta=target_qty - current_qty,
                has_wms_record=has_record,
                known_product=known_product,
                implicit_zero=False,
            )
        )

    # ── 补充盘点表未覆盖的仓库SKU（cover_uncovered=True 时视为0，否则跳过）──
    if not cover_uncovered:
        return lines

    not_in_filter = (
        InventoryRecord.product_jan.not_in(list(covered_jans))
        if covered_jans
        else True  # count sheet was empty — treat all WMS records as uncovered
    )
    uncovered_stmt = (
        select(InventoryRecord, Product)
        .join(Product, Product.jan_code == InventoryRecord.product_jan, isouter=True)
        .where(
            InventoryRecord.warehouse_id == warehouse_id,
            _customer_filter(customer_id),
            not_in_filter,
            InventoryRecord.quantity != 0,   # skip already-zero buckets
        )
        .order_by(InventoryRecord.product_jan.asc())
    )
    uncovered_rows = await session.execute(uncovered_stmt)

    for record, product in uncovered_rows.all():
        jan_code = record.product_jan
        current_qty = record.quantity

        delta_result = await session.scalar(
            select(func.coalesce(func.sum(StockTransaction.quantity_change), 0))
            .join(InventoryRecord, StockTransaction.inventory_record_id == InventoryRecord.id)
            .where(
                InventoryRecord.product_jan == jan_code,
                InventoryRecord.warehouse_id == warehouse_id,
                StockTransaction.created_at > cutoff,
            )
        )
        delta = int(delta_result or 0)
        target_qty = 0 + delta  # physical count = 0 for uncovered SKUs

        lines.append(
            CountDraftLine(
                jan_code=jan_code,
                product_name=product.name_jp if product else jan_code,
                count_quantity=0,
                delta_after_count=delta,
                target_quantity=target_qty,
                current_quantity=current_qty,
                adjust_delta=target_qty - current_qty,
                has_wms_record=True,
                known_product=product is not None,
                implicit_zero=True,
            )
        )

    return lines


def _customer_filter(customer_id: int | None):
    if customer_id is None:
        return InventoryRecord.customer_id.is_(None)
    return InventoryRecord.customer_id == customer_id


# ---------------------------------------------------------------------------
# Draft retrieval
# ---------------------------------------------------------------------------

async def get_inventory_count_draft(
    session: AsyncSession,
    draft_id: int,
    *,
    with_for_update: bool = False,
) -> InventoryCountDraft | None:
    stmt = select(InventoryCountDraft).where(InventoryCountDraft.id == draft_id)
    if with_for_update:
        stmt = stmt.with_for_update()
    return await session.scalar(stmt)


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------

async def apply_inventory_count_draft(
    session: AsyncSession,
    draft: InventoryCountDraft,
    user_id: int | None = None,
) -> InventoryCountApplyResult:
    if draft.status == "applied":
        return InventoryCountApplyResult(
            applied=False,
            adjusted_count=0,
            no_change_count=0,
            created_count=0,
            issues=["该草稿已经被应用过，不能重复提交"],
        )

    document = InventoryCountDocument.model_validate(draft.document)
    warehouse = await resolve_warehouse(session, document.warehouse_name)
    customer = None
    if document.customer_name:
        customer = await resolve_customer(session, document.customer_name)

    if warehouse is None:
        return InventoryCountApplyResult(
            applied=False,
            adjusted_count=0,
            no_change_count=0,
            created_count=0,
            issues=[f"仓库不存在: {document.warehouse_name}"],
        )

    adjusted = created = no_change = 0
    issues: list[str] = []
    customer_id = customer.id if customer else None

    for line in document.lines:
        if line.adjust_delta == 0:
            no_change += 1
            continue

        record = await session.scalar(
            select(InventoryRecord)
            .where(
                InventoryRecord.product_jan == line.jan_code,
                InventoryRecord.warehouse_id == warehouse.id,
                _customer_filter(customer_id),
            )
            .order_by(InventoryRecord.id.asc())
            .limit(1)
            .with_for_update()
        )

        if record is None:
            product_exists = bool(await session.scalar(
                select(Product.jan_code).where(Product.jan_code == line.jan_code).limit(1)
            ))
            if not product_exists:
                issues.append(f"跳过 {line.jan_code} ({line.product_name[:20]}): 商品字典中不存在")
                continue
            record = InventoryRecord(
                product_jan=line.jan_code,
                warehouse_id=warehouse.id,
                customer_id=customer_id,
                location_code="A-00-00",
                quantity=line.target_quantity,
            )
            session.add(record)
            await session.flush()
            created += 1
        else:
            record.quantity += line.adjust_delta
            await session.flush()
            adjusted += 1

        transaction = StockTransaction(
            inventory_record_id=record.id,
            transaction_type=StockTransactionType.adjust,
            quantity_change=line.adjust_delta,
            source="physical_count",
            note=(
                f"盘点日期:{document.count_date} "
                f"盘点量:{line.count_quantity} "
                f"盘后变动:{line.delta_after_count:+d}"
            ),
            user_id=user_id,
            transaction_date=document.count_date,
        )
        session.add(transaction)

    draft.status = "applied"
    await session.commit()

    return InventoryCountApplyResult(
        applied=True,
        adjusted_count=adjusted,
        no_change_count=no_change,
        created_count=created,
        issues=issues,
    )


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

def export_draft_to_excel(document: InventoryCountDocument) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "盘点对账"

    headers = [
        "JAN条码", "商品名称",
        f"盘点数量({document.count_date})",
        "盘后WMS变动", "目标库存", "当前WMS库存", "ADJUST量",
    ]
    ws.append(headers)

    for line in document.lines:
        ws.append([
            line.jan_code,
            line.product_name,
            line.count_quantity,
            line.delta_after_count,
            line.target_quantity,
            line.current_quantity,
            line.adjust_delta,
        ])

    # Column widths
    col_widths = [16, 40, 16, 14, 12, 12, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
