"""客户货量软预留服务。

流程：
1. 上传 Excel → `upsert_allocations_from_excel` → 解析行，UPSERT 到 CustomerAllocation，
   立即调 `_revalidate_for_jan` 评估哪些可以 reserved
2. 每次入库后 → `try_auto_reserve(session, jan_code, warehouse_id)` 触发同 JAN 的重新评估
3. 手动调转 → `try_reserve_one(session, allocation_id)`
4. 撤销 → `revert_to_waiting(session, allocation_id)` + 重新评估
5. 取消 → `cancel_allocation(session, allocation_id)`

防抽风核心：`_revalidate_for_jan` 持有 InventoryRecord 行锁（SELECT FOR UPDATE），
在同一事务内原子地重分配所有 reserved/waiting 行。
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime

import openpyxl
from sqlalchemy import and_, case, or_, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.common.customer_search import resolve_customer_names
from app.models.allocation_conflict_log import AllocationConflictLog
from app.models.customer_allocation import CustomerAllocation
from app.models.inventory_record import InventoryRecord
from app.models.product import Product
from app.models.warehouse import Warehouse
from app.schemas.inventory import (
    AllocationConflictLogRead,
    CustomerAllocationRead,
    CustomerAllocationStatusResult,
    CustomerAllocationUploadResult,
)
from app.services.product_alias import resolve_canonical_jan

ALLOCATION_WAREHOUSE_NAME = "普通仓库"


# ---------------------------------------------------------------------------
# 文件名日期解析
# ---------------------------------------------------------------------------

def _parse_date_from_filename(filename: str) -> date | None:
    """从文件名提取计划出库日期。

    优先匹配 YYYYMMDD，其次 YYYY-MM-DD / YYYY/MM/DD。
    """
    name = filename.replace("\\", "/").rsplit("/", 1)[-1]
    m = re.search(r"(\d{4})[/-]?(\d{2})[/-]?(\d{2})", name)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Excel 解析
# ---------------------------------------------------------------------------

def _parse_allocation_excel(
    content: bytes,
) -> list[tuple[str, str, int]]:
    """解析客户需求 Excel，返回 [(customer_name, jan_code, quantity), ...]。

    支持格式：
      - 每个 sheet 名即客户代码（mm / kk / cp 等）
      - 每行有 JAN（13位纯数字）和数量两列；列名不敏感，只要含 jan/JAN 和 数量/qty/quantity
      - 数量 ≤ 0 的行跳过
      - JAN 跨多行合并居中、数量分行填写时，合并单元格按锚点值回填，同 JAN 多行数量自动相加
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    rows: list[tuple[str, str, int]] = []

    for sheet in wb.worksheets:
        customer_name = sheet.title.strip()
        if not customer_name:
            continue

        # Vertically merged cells (e.g. one JAN spanning rows 33-34, centered) leave
        # value=None on every row except the top-left anchor; build a lookup so those
        # rows resolve to the anchor's value instead of being silently dropped.
        merged_value_map: dict[tuple[int, int], object] = {}
        for merged_range in sheet.merged_cells.ranges:
            anchor_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_value_map[(r, c)] = anchor_value

        # detect header row
        header_row_idx: int | None = None
        jan_col: int | None = None
        qty_col: int | None = None

        for row in sheet.iter_rows(max_row=10):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value).strip().lower()
                if "jan" in val or "条码" in val or "barcode" in val:
                    jan_col = cell.column
                    header_row_idx = cell.row
                if "数量" in val or "qty" in val or "quantity" in val or "個数" in val or "总数" in val:
                    qty_col = cell.column
            if jan_col and qty_col:
                break

        if not (jan_col and qty_col and header_row_idx):
            continue

        for row in sheet.iter_rows(min_row=header_row_idx + 1):
            jan_cell = row[jan_col - 1]
            qty_cell = row[qty_col - 1]
            jan_raw = jan_cell.value
            if jan_raw is None:
                jan_raw = merged_value_map.get((jan_cell.row, jan_cell.column))
            qty_raw = qty_cell.value
            if qty_raw is None:
                qty_raw = merged_value_map.get((qty_cell.row, qty_cell.column))
            if jan_raw is None or qty_raw is None:
                continue
            # Excel/Sheets often stores JAN as a number (e.g. 4987227028276.0);
            # str() on that float would append ".0", inflating the digit count to 14.
            if isinstance(jan_raw, float) and jan_raw.is_integer():
                jan_raw = int(jan_raw)
            jan = "".join(c for c in str(jan_raw) if c.isdigit())
            if len(jan) != 13:
                continue
            try:
                qty = int(float(qty_raw))
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue
            rows.append((customer_name, jan, qty))

    # Same (customer, jan) can legitimately appear on multiple rows — merged-cell
    # splits, or simply repeated line items — sum them instead of letting the
    # later UPSERT silently overwrite the earlier quantity.
    summed: dict[tuple[str, str], int] = {}
    for customer_name, jan, qty in rows:
        key = (customer_name, jan)
        summed[key] = summed.get(key, 0) + qty

    return [(customer_name, jan, qty) for (customer_name, jan), qty in summed.items()]


# ---------------------------------------------------------------------------
# 核心：重新评估某 JAN 在普通仓库的所有预留（持锁，原子）
# ---------------------------------------------------------------------------

async def _revalidate_for_jan(
    session: AsyncSession,
    jan_code: str,
    warehouse: Warehouse,
    exclude_ids: set[int] | None = None,
    trigger: str = "revalidate",
) -> None:
    """原子地重分配该 JAN 的所有 waiting/reserved 行。

    持有 InventoryRecord 行锁，按 planned_outbound_date ASC 贪心：
    累积量 ≤ 当前库存 → reserved；超出 → waiting。

    `exclude_ids`：本次重新评估时跳过这些行（不参与排队，也不占用库存额度）。
    用于手动撤销场景——撤销后若不排除自身，库存充足时会被立即重新判定为 reserved，
    导致撤销操作看起来毫无效果。

    任何 reserved → waiting 的降级（即库存不再能支撑某条原本已确认的预留，
    典型场景是该 JAN 的库存被预留系统之外的出库/调整消耗掉）都会写入
    `AllocationConflictLog`，`trigger` 标注是哪个动作触发了这次重新评估。
    """
    inv_record = await session.scalar(
        select(InventoryRecord).where(
            InventoryRecord.product_jan == jan_code,
            InventoryRecord.warehouse_id == warehouse.id,
        ).with_for_update()
    )
    current_stock = inv_record.quantity if inv_record else 0

    stmt = select(CustomerAllocation).where(
        CustomerAllocation.jan_code == jan_code,
        CustomerAllocation.status.in_(["waiting", "reserved"]),
    )
    if exclude_ids:
        stmt = stmt.where(CustomerAllocation.id.notin_(exclude_ids))
    stmt = stmt.order_by(
        CustomerAllocation.planned_outbound_date.asc(), CustomerAllocation.id.asc()
    ).with_for_update()
    allocs = (await session.scalars(stmt)).all()

    running = 0
    for alloc in allocs:
        running += alloc.quantity
        new_status = "reserved" if running <= current_stock else "waiting"
        if alloc.status == "reserved" and new_status == "waiting":
            session.add(AllocationConflictLog(
                jan_code=alloc.jan_code,
                customer_name=alloc.customer_name,
                planned_outbound_date=alloc.planned_outbound_date,
                quantity=alloc.quantity,
                current_stock=current_stock,
                trigger=trigger,
            ))
            from app.services.system_log import write_system_log  # lazy to avoid circular import
            await write_system_log(
                session,
                category="allocation_conflict",
                message=f"客户「{alloc.customer_name}」预留 {alloc.quantity} 件因库存不足（当前 {current_stock}）退回等待中",
                jan_code=alloc.jan_code,
                warehouse_name=warehouse.name,
            )
        alloc.status = new_status


# ---------------------------------------------------------------------------
# UPSERT：上传 Excel 时写入/更新预留行
# ---------------------------------------------------------------------------

async def upsert_allocations_from_excel(
    session: AsyncSession,
    content: bytes,
    filename: str,
    planned_outbound_date: date | None,
) -> CustomerAllocationUploadResult:
    if planned_outbound_date is None:
        planned_outbound_date = _parse_date_from_filename(filename)
    if planned_outbound_date is None:
        raise ValueError("无法从文件名解析出库日期，请在上传时手动指定日期")

    rows = _parse_allocation_excel(content)
    if not rows:
        raise ValueError("Excel 中未解析到有效行（需要 JAN 13位 + 数量列）")

    # 别名归一化：Excel 里仍可能用旧/别名JAN填报，统一解析到主JAN再写库存评估，
    # 否则预留行会卡在别名JAN上，永远查不到合并后的真实库存（见 product_alias.create_alias）。
    resolved: dict[tuple[str, str], int] = {}
    for customer_name, jan_code, quantity in rows:
        canonical_jan = await resolve_canonical_jan(session, jan_code)
        key = (customer_name, canonical_jan)
        resolved[key] = resolved.get(key, 0) + quantity
    rows = [(customer_name, jan_code, quantity) for (customer_name, jan_code), quantity in resolved.items()]

    warehouse = await session.scalar(
        select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
    )
    if warehouse is None:
        raise ValueError(f"找不到仓库「{ALLOCATION_WAREHOUSE_NAME}」，请先在系统中创建")

    stats = {"reserved": 0, "waiting": 0, "updated": 0, "skipped": 0}
    affected_jans: set[str] = set()

    for customer_name, jan_code, quantity in rows:
        # UPSERT: (date, customer, jan) 唯一。
        # - cancelled 行：无论数量是否相同都"复活"为 waiting（否则取消后无法通过重新上传找回，
        #   因为旧的 where 条件只比较数量，数量相同时整行被跳过、status 永远卡在 cancelled）。
        # - shipped 行：货已实际出库，重新上传不应改动，跳过。
        # - waiting/reserved 行：数量不变则跳过，数量变了才更新（status 不动，留给后续重新评估）。
        stmt = pg_insert(CustomerAllocation).values(
            planned_outbound_date=planned_outbound_date,
            customer_name=customer_name,
            jan_code=jan_code,
            quantity=quantity,
            status="waiting",
            source_filename=filename,
        ).on_conflict_do_update(
            constraint="uq_customer_alloc_date_customer_jan",
            set_={
                "quantity": quantity,
                "status": case(
                    (CustomerAllocation.status == "cancelled", "waiting"),
                    else_=CustomerAllocation.status,
                ),
                "source_filename": filename,
                "updated_at": datetime.utcnow(),
            },
            where=and_(
                CustomerAllocation.status != "shipped",
                or_(
                    CustomerAllocation.quantity != quantity,
                    CustomerAllocation.status == "cancelled",
                ),
            ),
        )
        result = await session.execute(stmt)
        if result.rowcount == 0:
            stats["skipped"] += 1
        else:
            stats["updated"] += 1
        affected_jans.add(jan_code)

    await session.flush()

    # 重新评估所有受影响的 JAN
    for jan_code in affected_jans:
        await _revalidate_for_jan(session, jan_code, warehouse, trigger="excel_upload")

    await session.flush()

    # 统计最终状态
    for customer_name, jan_code, _ in rows:
        if jan_code in affected_jans:
            pass  # counted below

    allocs = (await session.scalars(
        select(CustomerAllocation).where(
            CustomerAllocation.jan_code.in_(affected_jans),
            CustomerAllocation.planned_outbound_date == planned_outbound_date,
        )
    )).all()
    customer_names_in_upload = {r[0] for r in rows}
    for a in allocs:
        if a.customer_name not in customer_names_in_upload:
            continue
        if a.status == "reserved":
            stats["reserved"] += 1
        elif a.status == "waiting":
            stats["waiting"] += 1

    # subtract skipped/updated from reserved/waiting to avoid double-counting
    # Actually: recalculate properly
    stats["reserved"] = sum(1 for a in allocs if a.status == "reserved" and a.customer_name in customer_names_in_upload)
    stats["waiting"] = sum(1 for a in allocs if a.status == "waiting" and a.customer_name in customer_names_in_upload)

    await session.commit()

    return CustomerAllocationUploadResult(
        planned_outbound_date=planned_outbound_date,
        source_filename=filename,
        total_rows=len(rows),
        reserved=stats["reserved"],
        waiting=stats["waiting"],
        updated=stats["updated"],
        skipped=stats["skipped"],
    )


# ---------------------------------------------------------------------------
# 入库后自动调转钩子
# ---------------------------------------------------------------------------

async def try_auto_reserve(
    session: AsyncSession,
    jan_code: str,
    warehouse_id: int,
) -> None:
    """入库后调用，尝试将 waiting 行升级为 reserved。

    在调用方的事务内运行（不自行 commit），由调用方统一提交。
    只在仓库为普通仓库且存在 waiting 行时执行重分配。
    """
    warehouse = await session.get(Warehouse, warehouse_id)
    if warehouse is None or warehouse.name != ALLOCATION_WAREHOUSE_NAME:
        return

    has_waiting = await session.scalar(
        select(CustomerAllocation.id).where(
            CustomerAllocation.jan_code == jan_code,
            CustomerAllocation.status == "waiting",
        ).limit(1)
    )
    if has_waiting is None:
        return

    await _revalidate_for_jan(session, jan_code, warehouse, trigger="stock_in")


async def check_for_reservation_conflict(
    session: AsyncSession,
    jan_code: str,
    warehouse_id: int,
    trigger: str,
) -> None:
    """出库/调整库存后调用，检测该 JAN 是否有 reserved 预留因库存被消耗而不再满足。

    出库渠道（手动出库、秦丝同步、贸易出库、乐天订单等）和盘点调整都不知道
    `CustomerAllocation` 的存在，可能消耗掉本应留给某客户的库存。若不主动重新评估，
    该客户的预留会一直显示"已调转"，但实际库存已经不够——这里在每次出库/调整后
    重新核算，库存不足的 reserved 行会自动降级为 waiting，并写入冲突日志。

    在调用方的事务内运行（不自行 commit），由调用方统一提交。
    只在仓库为普通仓库且存在 reserved 行时执行重分配。
    """
    warehouse = await session.get(Warehouse, warehouse_id)
    if warehouse is None or warehouse.name != ALLOCATION_WAREHOUSE_NAME:
        return

    has_reserved = await session.scalar(
        select(CustomerAllocation.id).where(
            CustomerAllocation.jan_code == jan_code,
            CustomerAllocation.status == "reserved",
        ).limit(1)
    )
    if has_reserved is None:
        return

    await _revalidate_for_jan(session, jan_code, warehouse, trigger=trigger)


# ---------------------------------------------------------------------------
# 手动调转 / 撤销 / 取消 / 调整数量
# ---------------------------------------------------------------------------

async def update_allocation_quantity(
    session: AsyncSession,
    allocation_id: int,
    quantity: int,
) -> CustomerAllocation:
    """手动修正预留数量（用于纠正重复上传/录入错误导致的多算或少算）。

    修改后立即重新评估同 JAN 的所有预留，可能导致该行或其他行的 status 变化。
    """
    if quantity <= 0:
        raise ValueError("数量必须大于0；如需移除该条预留请使用「取消」")

    alloc = await session.scalar(
        select(CustomerAllocation).where(CustomerAllocation.id == allocation_id).with_for_update()
    )
    if alloc is None:
        raise ValueError("预留记录不存在")
    if alloc.status in ("shipped", "cancelled"):
        raise ValueError(f"当前状态 {alloc.status}，无法调整数量")

    alloc.quantity = quantity
    await session.flush()

    warehouse = await session.scalar(
        select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
    )
    if warehouse:
        await _revalidate_for_jan(session, alloc.jan_code, warehouse, trigger="manual_quantity_edit")

    await session.commit()
    await session.refresh(alloc)
    return alloc


async def try_reserve_one(
    session: AsyncSession,
    allocation_id: int,
) -> CustomerAllocation:
    """手动将一条 waiting 行调转为 reserved（若库存足够）。"""
    alloc = await session.scalar(
        select(CustomerAllocation).where(CustomerAllocation.id == allocation_id).with_for_update()
    )
    if alloc is None:
        raise ValueError("预留记录不存在")
    if alloc.status not in ("waiting",):
        raise ValueError(f"当前状态 {alloc.status}，无法调转")

    warehouse = await session.scalar(
        select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
    )
    if warehouse is None:
        raise ValueError(f"仓库「{ALLOCATION_WAREHOUSE_NAME}」不存在")

    await _revalidate_for_jan(session, alloc.jan_code, warehouse, trigger="manual_reserve")
    await session.flush()

    # Refresh to get updated status
    await session.refresh(alloc)
    if alloc.status == "waiting":
        raise ValueError("库存不足，无法调转")

    await session.commit()
    return alloc


async def revert_to_waiting(
    session: AsyncSession,
    allocation_id: int,
) -> CustomerAllocation:
    """撤销 reserved → waiting，并重新评估同 JAN 的其他 waiting 行。"""
    alloc = await session.scalar(
        select(CustomerAllocation).where(CustomerAllocation.id == allocation_id).with_for_update()
    )
    if alloc is None:
        raise ValueError("预留记录不存在")
    if alloc.status != "reserved":
        raise ValueError(f"当前状态 {alloc.status}，只有 reserved 状态可以撤销")

    alloc.status = "waiting"
    await session.flush()

    warehouse = await session.scalar(
        select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
    )
    if warehouse:
        # exclude self: 否则库存充足时会被立即重新判定为 reserved，撤销形同没做
        await _revalidate_for_jan(session, alloc.jan_code, warehouse, exclude_ids={alloc.id}, trigger="manual_revert")

    await session.commit()
    await session.refresh(alloc)
    return alloc


async def cancel_allocation(
    session: AsyncSession,
    allocation_id: int,
) -> CustomerAllocation:
    """取消预留（reserved/waiting → cancelled）。"""
    alloc = await session.scalar(
        select(CustomerAllocation).where(CustomerAllocation.id == allocation_id).with_for_update()
    )
    if alloc is None:
        raise ValueError("预留记录不存在")
    if alloc.status in ("shipped", "cancelled"):
        raise ValueError(f"当前状态 {alloc.status}，无法取消")

    prev_status = alloc.status
    alloc.status = "cancelled"
    await session.flush()

    # If was reserved, release capacity and re-evaluate others
    if prev_status == "reserved":
        warehouse = await session.scalar(
            select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
        )
        if warehouse:
            await _revalidate_for_jan(session, alloc.jan_code, warehouse, trigger="manual_cancel")

    await session.commit()
    await session.refresh(alloc)
    return alloc


async def mark_as_shipped(
    session: AsyncSession,
    allocation_id: int,
) -> CustomerAllocation:
    """手动标记该预留对应的货已实际出库（reserved/waiting → shipped）。

    出库渠道（秦丝同步/贸易出库等）的客户名格式与本表不一致，无法可靠自动匹配，
    因此由操作员在确认实际出库后手动标记，避免误匹配到错误客户、污染其他人的预留状态。
    标记后重新评估同 JAN 的其他行，确保基于出库后的真实库存重新核算（出库本身的
    库存扣减发生在对应的 stock_out 调用里，与本次状态标记是两个独立动作）。
    """
    alloc = await session.scalar(
        select(CustomerAllocation).where(CustomerAllocation.id == allocation_id).with_for_update()
    )
    if alloc is None:
        raise ValueError("预留记录不存在")
    if alloc.status in ("shipped", "cancelled"):
        raise ValueError(f"当前状态 {alloc.status}，无法标记已出库")

    alloc.status = "shipped"
    await session.flush()

    warehouse = await session.scalar(
        select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
    )
    if warehouse:
        await _revalidate_for_jan(session, alloc.jan_code, warehouse, trigger="mark_shipped")

    await session.commit()
    await session.refresh(alloc)
    return alloc


async def bulk_cancel_allocations(
    session: AsyncSession,
    customer_name: str,
    planned_outbound_date: date,
) -> int:
    """按客户名+计划出库日期批量取消（典型用途：上传时填错日期，整批撤销重传）。

    只取消 waiting/reserved 状态的行（shipped/cancelled 不受影响）。
    返回实际取消的行数。
    """
    resolved_name = await _resolve_single_customer_name(session, customer_name)
    allocs = (await session.scalars(
        select(CustomerAllocation).where(
            CustomerAllocation.customer_name == resolved_name,
            CustomerAllocation.planned_outbound_date == planned_outbound_date,
            CustomerAllocation.status.in_(["waiting", "reserved"]),
        ).with_for_update()
    )).all()

    if not allocs:
        return 0

    affected_jans: set[str] = set()
    had_reserved = False
    for alloc in allocs:
        if alloc.status == "reserved":
            had_reserved = True
        alloc.status = "cancelled"
        affected_jans.add(alloc.jan_code)

    await session.flush()

    if had_reserved:
        warehouse = await session.scalar(
            select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
        )
        if warehouse:
            for jan_code in affected_jans:
                await _revalidate_for_jan(session, jan_code, warehouse, trigger="bulk_cancel")

    await session.commit()
    return len(allocs)


async def bulk_mark_shipped_allocations(
    session: AsyncSession,
    customer_name: str,
    planned_outbound_date: date,
) -> int:
    """按客户名+计划出库日期一键标记已出库（典型用途：整批货已交给客户，逐条点太慢）。

    只标记 waiting/reserved 状态的行（shipped/cancelled 不受影响）。
    返回实际标记的行数。
    """
    resolved_name = await _resolve_single_customer_name(session, customer_name)
    allocs = (await session.scalars(
        select(CustomerAllocation).where(
            CustomerAllocation.customer_name == resolved_name,
            CustomerAllocation.planned_outbound_date == planned_outbound_date,
            CustomerAllocation.status.in_(["waiting", "reserved"]),
        ).with_for_update()
    )).all()

    if not allocs:
        return 0

    affected_jans: set[str] = set()
    for alloc in allocs:
        alloc.status = "shipped"
        affected_jans.add(alloc.jan_code)

    await session.flush()

    warehouse = await session.scalar(
        select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
    )
    if warehouse:
        for jan_code in affected_jans:
            await _revalidate_for_jan(session, jan_code, warehouse, trigger="mark_shipped")

    await session.commit()
    return len(allocs)


# ---------------------------------------------------------------------------
# 查询：货齐了状态
# ---------------------------------------------------------------------------

async def _fuzzy_customer_names(session: AsyncSession, query: str) -> list[str]:
    """把模糊输入解析为命中的客户名列表（子串/简繁体/拼音首字母，见 app.common.customer_search）。"""
    candidates = (await session.scalars(select(CustomerAllocation.customer_name).distinct())).all()
    return resolve_customer_names(query, candidates)


async def _resolve_single_customer_name(session: AsyncSession, query: str) -> str:
    """把模糊输入解析为唯一客户名，供会修改数据的操作使用（批量取消/批量标记出库）。

    命中 0 个或多个客户名都报错——绝不在批量修改场景里猜测，避免误打误撞影响到无关客户。
    """
    matched = await _fuzzy_customer_names(session, query)
    if not matched:
        raise ValueError(f"未找到匹配「{query}」的客户")
    if len(matched) > 1:
        raise ValueError(f"「{query}」匹配到多个客户：{'、'.join(matched)}，请输入更精确的关键字")
    return matched[0]


async def get_allocation_status(
    session: AsyncSession,
    customer_name: str | None = None,
    planned_outbound_date: date | None = None,
    status_filter: str | None = None,
    jan_code: str | None = None,
) -> list[CustomerAllocationRead]:
    """查询预留状态，实时附带当前库存和商品名。

    `customer_name` 为模糊匹配（子串/简繁体互认/拼音首字母，见 app.common.customer_search），
    可能命中多个客户——这是只读查询，不像批量操作那样需要收窄到唯一客户。
    """
    warehouse = await session.scalar(
        select(Warehouse).where(Warehouse.name == ALLOCATION_WAREHOUSE_NAME)
    )

    stmt = select(CustomerAllocation)
    if customer_name:
        matched_names = await _fuzzy_customer_names(session, customer_name)
        if not matched_names:
            return []
        stmt = stmt.where(CustomerAllocation.customer_name.in_(matched_names))
    if planned_outbound_date:
        stmt = stmt.where(CustomerAllocation.planned_outbound_date == planned_outbound_date)
    if status_filter:
        stmt = stmt.where(CustomerAllocation.status == status_filter)
    if jan_code:
        stmt = stmt.where(CustomerAllocation.jan_code == jan_code)
    stmt = stmt.order_by(
        CustomerAllocation.planned_outbound_date.asc(),
        CustomerAllocation.customer_name.asc(),
        CustomerAllocation.jan_code.asc(),
    )

    allocs = (await session.scalars(stmt)).all()
    if not allocs:
        return []

    jan_codes = list({a.jan_code for a in allocs})

    products_map: dict[str, str] = {}
    for p in (await session.scalars(select(Product).where(Product.jan_code.in_(jan_codes)))).all():
        products_map[p.jan_code] = p.name_jp

    stock_map: dict[str, int] = {}
    if warehouse:
        for rec in (await session.scalars(
            select(InventoryRecord).where(
                InventoryRecord.product_jan.in_(jan_codes),
                InventoryRecord.warehouse_id == warehouse.id,
            )
        )).all():
            stock_map[rec.product_jan] = stock_map.get(rec.product_jan, 0) + rec.quantity

    result = []
    for a in allocs:
        r = CustomerAllocationRead.model_validate(a)
        r.product_name = products_map.get(a.jan_code)
        r.current_stock = stock_map.get(a.jan_code)
        result.append(r)
    return result


async def get_allocation_summary(
    session: AsyncSession,
    customer_name: str,
    planned_outbound_date: date,
) -> CustomerAllocationStatusResult:
    """查看某客户某日期的货齐了状态汇总。`customer_name` 模糊匹配但要求唯一命中
    （结果用一个 customer_name 标签展示总览，命中多个客户会产生误导，因此报错而非合并）。
    """
    resolved_name = await _resolve_single_customer_name(session, customer_name)
    items = await get_allocation_status(session, customer_name=resolved_name, planned_outbound_date=planned_outbound_date)
    return CustomerAllocationStatusResult(
        customer_name=resolved_name,
        planned_outbound_date=planned_outbound_date,
        ready_count=sum(1 for i in items if i.status == "reserved"),
        waiting_count=sum(1 for i in items if i.status == "waiting"),
        shipped_count=sum(1 for i in items if i.status == "shipped"),
        items=items,
    )


async def get_daily_allocation_overview(
    session: AsyncSession,
    planned_outbound_date: date,
) -> list[CustomerAllocationStatusResult]:
    """某天所有客户的预留情况一览，按客户名分组（装柜前一次性检查还缺哪些商品）。"""
    items = await get_allocation_status(session, planned_outbound_date=planned_outbound_date)

    by_customer: dict[str, list[CustomerAllocationRead]] = {}
    for item in items:
        by_customer.setdefault(item.customer_name, []).append(item)

    results = []
    for customer_name in sorted(by_customer):
        customer_items = by_customer[customer_name]
        results.append(CustomerAllocationStatusResult(
            customer_name=customer_name,
            planned_outbound_date=planned_outbound_date,
            ready_count=sum(1 for i in customer_items if i.status == "reserved"),
            waiting_count=sum(1 for i in customer_items if i.status == "waiting"),
            shipped_count=sum(1 for i in customer_items if i.status == "shipped"),
            items=customer_items,
        ))
    return results


# ---------------------------------------------------------------------------
# 查询：预留冲突日志
# ---------------------------------------------------------------------------

async def get_conflict_logs(
    session: AsyncSession,
    jan_code: str | None = None,
    customer_name: str | None = None,
    limit: int = 200,
) -> list[AllocationConflictLogRead]:
    """查询预留冲突日志（reserved→waiting 的降级事件），按时间倒序。"""
    stmt = select(AllocationConflictLog)
    if jan_code:
        stmt = stmt.where(AllocationConflictLog.jan_code == jan_code)
    if customer_name:
        stmt = stmt.where(AllocationConflictLog.customer_name == customer_name)
    stmt = stmt.order_by(AllocationConflictLog.created_at.desc()).limit(limit)

    logs = (await session.scalars(stmt)).all()
    if not logs:
        return []

    jan_codes = list({log.jan_code for log in logs})
    products_map: dict[str, str] = {}
    for p in (await session.scalars(select(Product).where(Product.jan_code.in_(jan_codes)))).all():
        products_map[p.jan_code] = p.name_jp

    result = []
    for log in logs:
        r = AllocationConflictLogRead.model_validate(log)
        r.product_name = products_map.get(log.jan_code)
        result.append(r)
    return result
