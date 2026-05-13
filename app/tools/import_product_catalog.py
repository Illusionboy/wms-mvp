import argparse
import asyncio
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.init_db import init_db
from app.db.session import AsyncSessionLocal, engine
from app.models.product import Product


@dataclass(frozen=True)
class ProductCatalogRow:
    jan_code: str
    name_jp: str
    name_zh: str | None
    units_per_case: int | None


async def main() -> None:
    args = _parse_args()
    await init_db()
    async with AsyncSessionLocal() as session:
        result = await import_product_catalog(
            session=session,
            file_path=Path(args.file),
            dry_run=args.dry_run,
        )
    await engine.dispose()
    action = "Would import" if args.dry_run else "Imported"
    print(
        f"{action} product catalog: "
        f"{result['created']} created, {result['updated']} updated, {result['skipped']} skipped"
    )


async def import_product_catalog(
    session: AsyncSession,
    file_path: Path,
    dry_run: bool = False,
) -> dict[str, int]:
    rows = _read_product_catalog_rows(file_path)
    created = 0
    updated = 0

    for row in rows:
        product = await session.get(Product, row.jan_code)
        if product is None:
            created += 1
            if not dry_run:
                session.add(
                    Product(
                        jan_code=row.jan_code,
                        name_jp=row.name_jp,
                        name_zh=row.name_zh,
                        units_per_case=row.units_per_case,
                    )
                )
            continue

        changed = False
        if row.name_jp and product.name_jp != row.name_jp:
            product.name_jp = row.name_jp
            changed = True
        if row.name_zh and product.name_zh != row.name_zh:
            product.name_zh = row.name_zh
            changed = True
        if row.units_per_case is not None and product.units_per_case != row.units_per_case:
            product.units_per_case = row.units_per_case
            changed = True
        if changed:
            updated += 1

    if not dry_run:
        await session.commit()
    return {"created": created, "updated": updated, "skipped": len(rows) - created - updated}


def _read_product_catalog_rows(file_path: Path) -> list[ProductCatalogRow]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    column_indexes = _catalog_column_indexes(header)

    rows_by_jan: dict[str, ProductCatalogRow] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        jan_code = _normalize_jan(row[column_indexes["jan"]] if len(row) > column_indexes["jan"] else None)
        if not jan_code:
            continue
        name_jp = _clean_text(row[column_indexes["name_jp"]] if len(row) > column_indexes["name_jp"] else None)
        name_zh = _clean_text(row[column_indexes["name_zh"]] if len(row) > column_indexes["name_zh"] else None) or None
        units_per_case = _normalize_positive_int(
            row[column_indexes["units_per_case"]] if len(row) > column_indexes["units_per_case"] else None
        )
        if not name_jp:
            name_jp = jan_code
        rows_by_jan[jan_code] = ProductCatalogRow(
            jan_code=jan_code,
            name_jp=name_jp[:255],
            name_zh=name_zh[:255] if name_zh else None,
            units_per_case=units_per_case,
        )
    workbook.close()
    return list(rows_by_jan.values())


def _catalog_column_indexes(header: tuple[object, ...]) -> dict[str, int]:
    normalized_header = {_clean_text(value): index for index, value in enumerate(header)}
    required_columns = {
        "jan": ["JAN", "jan"],
        "name_jp": ["商品名", "日文", "name_jp"],
        "name_zh": ["中文", "name_zh"],
        "units_per_case": ["箱入れ数", "箱入数", "箱入", "units_per_case"],
    }
    indexes: dict[str, int] = {}
    for key, names in required_columns.items():
        for name in names:
            if name in normalized_header:
                indexes[key] = normalized_header[name]
                break
        if key not in indexes:
            raise ValueError(f"Missing required product catalog column: {names[0]}")
    return indexes


def _normalize_jan(value: object) -> str:
    text = _clean_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return "".join(ch for ch in text if ch.isdigit())


def _normalize_positive_int(value: object) -> int | None:
    text = _clean_text(value)
    if not text:
        return None
    try:
        parsed = int(float(text))
    except ValueError:
        return None
    if parsed <= 0:
        return None
    return parsed


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import product master catalog from Excel.")
    parser.add_argument("--file", required=True, help="Excel file. Required columns: JAN, 商品名, 中文, 箱入れ数.")
    parser.add_argument("--dry-run", action="store_true", help="Read and count changes without writing to database.")
    return parser.parse_args()


if __name__ == "__main__":
    asyncio.run(main())
