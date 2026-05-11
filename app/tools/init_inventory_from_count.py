import argparse
import asyncio
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.init_db import init_db
from app.db.session import AsyncSessionLocal, engine
from app.models.customer import Customer
from app.models.inventory_record import InventoryRecord
from app.models.product import Product
from app.models.warehouse import Warehouse


DEFAULT_LOCATION_CODE = "A-00-00"


@dataclass(frozen=True)
class CountRow:
    jan_code: str
    quantity: int


async def main() -> None:
    args = _parse_args()
    await init_db()
    async with AsyncSessionLocal() as session:
        rakuten_count = await import_count_file(
            session=session,
            file_path=Path(args.rakuten_file),
            warehouse_name="乐天仓库",
            customer_name="乐天",
            replace=args.replace,
        )
        normal_count = await import_count_file(
            session=session,
            file_path=Path(args.normal_file),
            warehouse_name="普通仓库",
            customer_name="店铺",
            replace=args.replace,
        )
    await engine.dispose()
    print(f"Imported 乐天仓库/乐天: {rakuten_count} rows")
    print(f"Imported 普通仓库/店铺: {normal_count} rows")


async def import_count_file(
    session: AsyncSession,
    file_path: Path,
    warehouse_name: str,
    customer_name: str,
    replace: bool,
) -> int:
    rows = _read_count_rows(file_path)
    warehouse = await _get_or_create_warehouse(session, warehouse_name)
    customer = await _get_or_create_customer(session, customer_name)

    if replace:
        await session.execute(
            delete(InventoryRecord).where(
                InventoryRecord.warehouse_id == warehouse.id,
                InventoryRecord.customer_id == customer.id,
            )
        )
        await session.flush()

    imported_count = 0
    for row in rows:
        product = await _get_or_create_product(session, row.jan_code)
        statement = select(InventoryRecord).where(
            InventoryRecord.product_jan == product.jan_code,
            InventoryRecord.warehouse_id == warehouse.id,
            InventoryRecord.customer_id == customer.id,
            InventoryRecord.location_code == DEFAULT_LOCATION_CODE,
            InventoryRecord.expiration_date.is_(None),
        )
        record = await session.scalar(statement)
        if record is None:
            record = InventoryRecord(
                product_jan=product.jan_code,
                warehouse_id=warehouse.id,
                customer_id=customer.id,
                location_code=DEFAULT_LOCATION_CODE,
                quantity=row.quantity,
            )
            session.add(record)
        else:
            record.quantity = row.quantity
        imported_count += 1

    await session.commit()
    return imported_count


def _read_count_rows(file_path: Path) -> list[CountRow]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows: list[CountRow] = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        jan_code = _normalize_jan(row[0] if len(row) > 0 else None)
        quantity = _normalize_quantity(row[1] if len(row) > 1 else None)
        if not jan_code or quantity is None:
            continue
        rows.append(CountRow(jan_code=jan_code, quantity=quantity))
    workbook.close()
    return rows


def _normalize_jan(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        value = int(value)
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return "".join(ch for ch in text if ch.isdigit())


def _normalize_quantity(value: object) -> int | None:
    if value is None:
        return None
    try:
        quantity = int(float(str(value).strip()))
    except ValueError:
        return None
    if quantity < 0:
        return None
    return quantity


async def _get_or_create_warehouse(session: AsyncSession, name: str) -> Warehouse:
    warehouse = await session.scalar(select(Warehouse).where(Warehouse.name == name))
    if warehouse is not None:
        return warehouse
    warehouse = Warehouse(name=name)
    session.add(warehouse)
    await session.flush()
    return warehouse


async def _get_or_create_customer(session: AsyncSession, name: str) -> Customer:
    customer = await session.scalar(select(Customer).where(Customer.name == name))
    if customer is not None:
        return customer
    customer = Customer(name=name)
    session.add(customer)
    await session.flush()
    return customer


async def _get_or_create_product(session: AsyncSession, jan_code: str) -> Product:
    product = await session.get(Product, jan_code)
    if product is not None:
        return product
    product = Product(jan_code=jan_code, name_jp=jan_code)
    session.add(product)
    await session.flush()
    return product


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Initialize inventory from two no-header Excel count files.")
    parser.add_argument("--rakuten-file", required=True, help="Excel file for 乐天仓库. Column A=JAN, Column B=quantity.")
    parser.add_argument("--normal-file", required=True, help="Excel file for 普通仓库. Column A=JAN, Column B=quantity.")
    parser.add_argument("--replace", action="store_true", help="Replace existing records for these warehouse/customer pairs.")
    return parser.parse_args()


if __name__ == "__main__":
    asyncio.run(main())
